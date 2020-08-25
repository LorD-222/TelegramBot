import config
import telebot
from telebot import apihelper
import urllib.request 
import pyodbc
from datetime import datetime
import openpyxl

# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
#wb.create_sheet(title = 'Первый лист', index = 0)

# получаем лист, с которым будем работать
sheet = wb.active

driver = 'DRIVER={SQL Server}'
server = 'SERVER=dc-sql12-db\\db'
db = 'DATABASE=Inventory'
conn_str = ';'.join([driver, server, db])
 
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()
 
cursor.execute("SELECT hostname FROM dbo.main WHERE hostname = '90514-ws404' ")
row = cursor.fetchone()
rest_of_rows = cursor.fetchall()
#for item in rest_of_rows:
c1 = sheet.cell(row = 1, column = 1)
print(row)
c1.value = row
#   print(item)
wb.save('documents\\example.xlsx')